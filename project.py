import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import random



url = "https://codeforces.com/problemset"
response = requests.get(url)
soup = BeautifulSoup(response.text, 'html.parser')

options = soup.find_all('option')

#Problēmu tipu saglabāšana vārdnīcās (codeforces.com/problemset add tags)
tag_names = {}
success_rate = {}
for option in options:
    tag = option.get('value')
    if tag != "" and tag != "combine-tags-by-or":
        tag_names[tag] = 0
        success_rate[tag] = 0
    

url_beginning = "https://codeforces.com/profile/"
username = input("Provide your codeforces USERNAME or FULL LINK: ")

#profila pilnā linka iegūšana
if url_beginning not in username:
    username = url_beginning + username

name = username[len(url_beginning):]

response = requests.get(username)
actual_url = response.url

if actual_url != username:
    print("The name you provided is not correct or does not exist! Please run the program again and provide a valid username.")
    exit()

soup = BeautifulSoup(response.text, 'html.parser')
div_container = soup.find(class_="info")
li_elements = div_container.find_all('span', style = "font-weight:bold;")

#reitinga meklēšana
rating = 0
if len(li_elements) > 0:
    rating = li_elements[0].text

rating = int(rating)
rating_lowerbound = min(3000, max(800, rating - 200 - rating%100))
rating_upperbound = min(3500, max(1200, rating + 400 - rating%100))


all_problems = set()
encountered_problems = []
full_tag_count = tag_names.copy()
cur_page = 1

total_difficulty = 0
difficulty_count = 0
successful_solves = 0

#visu problēmu lapu izskatīšana
while True:

    cur_url =  "https://codeforces.com/submissions/" + name + "/page/" + str(cur_page)
    response = requests.get(cur_url)


    soup = BeautifulSoup(response.text, 'html.parser')

    submission_data = soup.find_all("tr")
    sub_count = 0
    for row in submission_data:
        items = row.find_all("td")
        if len(items) == 0:
            continue
        
        #problēmas rezultāta apskate
        accepted = False
        if items[5].find("span", class_="verdict-accepted"):
            accepted = True
        
        #katras problēmas tipa saskaitīšana, saglabāšana
        href = items[3].find("a").get("href")
        build_link = "https://codeforces.com" + href
        response = requests.get(build_link)
        soup = BeautifulSoup(response.text, 'html.parser')
        tags = soup.find_all("span", class_="tag-box")
        for tag in tags:
            parsed = tag.text.replace("\n", "").replace("  ", "")
            filtered = [x for x in tag_names.keys() if x in parsed]
            for x in filtered:
                full_tag_count[x] += 1
                if build_link not in encountered_problems:
                    tag_names[x] += 1
                if accepted:
                    success_rate[x] += 1

      
        if accepted:
            all_problems.add(build_link)

        encountered_problems.append(build_link)
        sub_count += 1
   
    if sub_count < 50:
        break
    
    cur_page += 1



avg_rate = 0
rate_count = 0
avg_tag_frequency = 0

#katras problēmas tipa, kopējo risinājumu vidējo vērtību aprēķināšana
#piemēram - kopēji apskatītas 10 matemātikas problēmas, taču pareizi atrisinātas tikai 6 --> 60% precizitāte problēmu tipam - matemātika.

for key in tag_names:
    avg_tag_frequency += tag_names[key]
    if full_tag_count[key] == 0:
        continue
    success_rate[key] = round(success_rate[key]/full_tag_count[key]*100)
    avg_rate += success_rate[key]
    rate_count += 1

avg_rate /= rate_count
avg_tag_frequency /= len(encountered_problems)

success_perc = 0
if len(encountered_problems) != 0:
    success_perc = round(len(all_problems)/len(encountered_problems)*100)


#reitinga diapazaonas problēmu meklēšana
problem_url = "https://codeforces.com/problemset/?tags={}-{}".format(rating_lowerbound, rating_upperbound)
response = requests.get(problem_url)
soup = BeautifulSoup(response.text, 'html.parser')

last_page = int(soup.find_all("span", class_="page-index")[-1].text)
recommended_problems = []


valid_length = True
visited_pages = set()

while valid_length:
    page = random.randint(1, last_page)

    #lai nemeklētu bezgalīgi, ja gadījumā dotajā diapazonā nav ko atrisināt
    if page in visited_pages:
        if len(visited_pages) == last_page:
            break
        continue

    visited_pages.add(page)
    problem_url = "https://codeforces.com/problemset/page/{}?tags={}-{}".format(page, rating_lowerbound, rating_upperbound)
    response = requests.get(problem_url)
    soup = BeautifulSoup(response.text, 'html.parser')

    td_con = soup.find_all("td")
    for div in td_con:
        for elements in div.find_all("div", style="float: right; font-size: 1.1rem; padding-top: 1px; text-align: right;"):
            for li in elements.find_all("a"):
                if li.text == "*special problem":
                    continue
                
                #pievienos problēmu tikai, ja tās tipa precizitāte ir zem vidējā, vai ja konkrētais algoritms netiek bieži apskatīts
                #piemēram - vidēja kopējā veiksmīgu iesniegumu precizitāte: 56%, 'data structures' algoritmu precizitāte: 40% --> 'data structure' uzdevums tiks pievienots.
                #piemēram - vidēji no katra tipa atrisinātas 10 problēmas, bet no 'shortest paths' tikai 7 problēmas --> 'shortest paths' uzdevums tiks pievienots.
                
                if success_rate[li.text] <= avg_rate or tag_names[li.text] <= avg_tag_frequency:
                    if len(recommended_problems) < 15:
                        problem_id = div.find("div", style="float: left;").find("a").get("href")
                        build_link = "https://codeforces.com" + problem_id
                        if build_link not in all_problems and build_link not in recommended_problems and len(recommended_problems) < 15:
                            recommended_problems.append(build_link)
                        if len(recommended_problems) == 15:
                            valid_length = False
                            break

                if not valid_length:
                    break

        if not valid_length:
            break

# Datu sakārtošana priekš attēlošanas excelii (visvairāk/vismazāk apskatītie problēmu tipi)
        
sorted_tags = dict(sorted(tag_names.items(), key=lambda item: item[1], reverse=True))
sorted_tags2 = dict(sorted(tag_names.items(), key=lambda item: item[1], reverse=False))
first_10_tags = [key for key in list(sorted_tags.keys())[:10]]
first_10_values = [sorted_tags[key] for key in list(sorted_tags.keys())[:10]]
last_5_tags = [key for key in list(sorted_tags2.keys())[:5]]
last_5_values = [sorted_tags2[key] for key in list(sorted_tags2.keys())[:5]]


#datu apstrāde, ierakstīšana excel failā
wb = Workbook()
ws = wb.active

problemdata = [first_10_tags, first_10_values, [success_rate[x] for x in first_10_tags]]
problemdata2 = [last_5_tags, last_5_values, [success_rate[x] for x in last_5_tags]]
profiledata = [['USERNAME', username], ['RATING', rating], ['SOLVED PROBLEMS', len(all_problems)], ['SUBMISSION SUCCESS RATE', success_perc]]

for col_letter in ['A', 'B', 'C', 'D']:
    col_index = ord(col_letter) - ord('A') + 1
    ws.column_dimensions[col_letter].width = 40

ws.row_dimensions[1].height = 30
ws.row_dimensions[4].height = 20
ws.row_dimensions[11].height = 20 

orange_border = Border(
    left=Side(style='thin', color='FFA500'),
    right=Side(style='thin', color='FFA500'),
    top=Side(style='thin', color='FFA500'),
    bottom=Side(style='thin', color='FFA500')
)
orange_fill = PatternFill(start_color="FFD7B5", end_color="FFD7B5", fill_type="solid")

for i in range(1, 3):
    for j in range(1, 5):
        cell = ws.cell(row=i, column=j, value = profiledata[j-1][i-1])
        if i == 1:
            cell.font = Font(bold=True)
        elif j==1:
            cell.value = name
            cell.hyperlink = profiledata[j-1][i-1]
            cell.style = 'Hyperlink' 
            cell.font = Font(color="964B00", bold=True)
        elif j==4:
            cell.value = str(profiledata[j-1][i-1]) + " %"

        if i != 1:
            cell.fill = orange_fill
    
        cell.alignment = Alignment(horizontal='left')
        cell.border = orange_border


ws.cell(row=4, column=1, value = "Recommended Problems")
ws.cell(row=4, column=1).font = Font(bold=True)
ws.cell(row=4, column=1).border = orange_border

for i in range(5, min(20, 5+len(recommended_problems))):
    cell = ws.cell(row=i,column=1)
    cell.value = recommended_problems[i-5].replace('https://codeforces.com/problemset/', '')
    cell.hyperlink = recommended_problems[i-5]
    cell.style = 'Hyperlink'
    cell.border = orange_border
    cell.fill = orange_fill

titles = ['Frequently Solved Problems', 'Problem Count', 'Success Rate']
titles2 = ['Least Encountered Problems', 'Problem Count', 'Success Rate']

for i in range(21, 32):
    for j in range(1, 4):
        cell = ws.cell(row=i,column=j)

        if i == 21:
            cell.value = titles[j-1]
            cell.font = Font(bold=True)
        elif j==3:
            cell.value = str(problemdata[j-1][i-22]) + " %"
        else:
            cell.value = problemdata[j-1][i-22]

        if i != 21:
            cell.fill = orange_fill

        cell.alignment = Alignment(horizontal='left')
        cell.border = orange_border

for i in range(33, 38):
    for j in range(1, 4):
        cell = ws.cell(row=i,column=j)

        if i == 33:
            cell.value = titles2[j-1]
            cell.font = Font(bold=True)
        elif j==3:
            cell.value = str(problemdata2[j-1][i-34]) + " %"
        else:
            cell.value = problemdata2[j-1][i-34]

        if i != 33:
            cell.fill = orange_fill

        cell.alignment = Alignment(horizontal='left')
        cell.border = orange_border


wb.save('user_statistics.xlsx')